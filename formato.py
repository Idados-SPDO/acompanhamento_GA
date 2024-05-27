
from openpyxl.styles import PatternFill, Font, Border, Side
 
def formatar_excel(writer, sheet_name, df):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(border_style="thin"), 
                           bottom=Side(border_style="thin"), 
                           left=Side(border_style="thin"), 
                           right=Side(border_style="thin"))

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border

    data_border = Border(top=Side(border_style="thin"), 
                         bottom=Side(border_style="thin"), 
                         left=Side(border_style="thin"), 
                         right=Side(border_style="thin"))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = data_border

    ws.sheet_view.showGridLines = False

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

