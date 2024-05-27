import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


# Functions
def atualizar_base(base, consulta):
    base_atualizada = base.copy()
    
    for index, row in consulta.iterrows():
        solicitacao = row['Solicitação']
        if solicitacao in base['Solicitação'].values:
            for col in consulta.columns:
                if col in base.columns:
                    if base.loc[base['Solicitação'] == solicitacao, col].values[0] != row[col]:
                        base_atualizada.loc[base_atualizada['Solicitação'] == solicitacao, col] = row[col]
        else:
            base_atualizada = base_atualizada.append(row, ignore_index=True)
    
    return base_atualizada

def calcular_status(row):
    data_sistema = pd.Timestamp.now()
    diferenca = (data_sistema - row['Data Retorno']).days
    if row['Data Retorno'] > data_sistema:
        return 'No Prazo'
    elif diferenca <= 5:
        return 'Px Prazo Final'
    else:
        return 'Em Atraso'
    
def formatar_excel(writer, sheet_name, df):
    # Adiciona o DataFrame ao writer
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Acessa a planilha do writer
    ws = writer.sheets[sheet_name]
    
    # Estilo para os cabeçalhos
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")  # Azul claro
    header_font = Font(color="000000", bold=True)  # Cor preta para o texto dos cabeçalhos
    header_border = Border(top=Side(border_style="thin"), 
                           bottom=Side(border_style="thin"), 
                           left=Side(border_style="thin"), 
                           right=Side(border_style="thin"))  # Borda completa para os cabeçalhos
    
    # Aplica o estilo aos cabeçalhos
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
    
    # Estilo para os dados
    data_border = Border(top=Side(border_style="thin"), 
                         bottom=Side(border_style="thin"), 
                         left=Side(border_style="thin"), 
                         right=Side(border_style="thin"))  # Borda completa para os dados
    
    # Aplica o estilo aos dados
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = data_border
    
    # Oculta as linhas de grade na planilha
    ws.sheet_view.showGridLines = False

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Fator de ajuste para garantir que o texto não seja cortado
        ws.column_dimensions[column].width = adjusted_width

def main():

    construcao = pd.read_excel('Acompanhamento_GA.xlsx', index_col=False, sheet_name='Construção')
    industria = pd.read_excel('Acompanhamento_GA.xlsx', index_col=False, sheet_name='Indústria')

    construcao['Data Solicitação'] = pd.to_datetime(construcao['Data Solicitação'])
    industria['Data Solicitação'] = pd.to_datetime(industria['Data Solicitação'])
    construcao['Data Solicitação'] = construcao['Data Solicitação'].dt.date
    industria['Data Solicitação'] = industria['Data Solicitação'].dt.date

    consulta = pd.read_excel('query.xlsx', index_col=False)
    consulta = consulta[['Solicitação n°', 'Status da Atividade', 'Solicitante', 'Executor', 
                            'Tipo de solicitação', 'JOB/Serviço', 'Prazo p/retorno', 'Solicitado em', 'Solicitado por']]

    consulta = consulta.rename(columns={
        'Solicitação n°': 'Solicitação',
        'Status da Atividade': 'Status do GA',
        'Solicitante': 'Área',
        'Executor': 'Executor',
        'Tipo de solicitação': 'Tipo de Solicitação',
        'JOB/Serviço': 'JOB',
        'Prazo p/retorno': 'Data Retorno',
        'Solicitado em': 'Data Solicitação',
        'Solicitado por': 'Solicitante'
    })

    consulta['Data Solicitação'] = pd.to_datetime(consulta['Data Solicitação'])
    consulta['Data Solicitação'] = consulta['Data Solicitação'].dt.date

    consulta_construcao = consulta[consulta['Área'].str.contains('Construção', case=False, na=False)]
    consulta_industria = consulta[consulta['Área'].str.contains('Indústria', case=False, na=False)]

    construcao_atualizada = atualizar_base(construcao, consulta_construcao)
    industria_atualizada = atualizar_base(industria, consulta_industria)

    construcao_atualizada['Status da Demanda'] = construcao_atualizada.apply(calcular_status, axis=1)
    industria_atualizada['Status da Demanda'] = industria_atualizada.apply(calcular_status, axis=1)

    construcao_atualizada['Status da Demanda'] = construcao_atualizada.apply(lambda row: 'Finalizada' if row['Status do GA'] in ['Concluída', 'Cancelada'] else row['Status da Demanda'] , axis=1)
    industria_atualizada['Status da Demanda'] = industria_atualizada.apply(lambda row: 'Finalizada' if row['Status do GA'] in ['Concluída', 'Cancelada'] else row['Status da Demanda'] , axis=1)

    construcao_atualizada['Data Retorno'] = construcao_atualizada['Data Retorno'].dt.date
    industria_atualizada['Data Retorno'] = industria_atualizada['Data Retorno'].dt.date

    with pd.ExcelWriter('Acompanhamento_GA.xlsx', engine='openpyxl') as writer:
        formatar_excel(writer, 'Construção', construcao_atualizada)
        formatar_excel(writer, 'Indústria', industria_atualizada)


if __name__ == "__main__":
    main()